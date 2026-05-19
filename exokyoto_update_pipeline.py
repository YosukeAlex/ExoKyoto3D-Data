#!/usr/bin/env python3
"""
exokyoto_update_pipeline.py — Unified ExoKyotoDataF update pipeline.

End-to-end flow for adding new entries / regenerating distribution artifacts:

  1. Diff new xlsx vs previous (find added/removed/changed planets)
  2. (optional) Query NASA Exoplanet Archive to fill NASA cols for added rows
  3. (optional) Query ADS for paper references for added rows
  4. Run duplicate sanity check
  5. xlsx -> CSV -> EKDBIN1 .bin
  6. Archive previous bin to data/past/ExoKyotoDataF<YYYYMMDDNN>.bin
  7. Install new bin to data/latest/ExoKyotoDataF.bin
  8. Update local version.json (data_version, release_date, notes)
  9. Print git commands for the user to push to ExoKyoto3D-Data repo

Layout assumed:
  gaia/ExoKyotoData/data/latest/ExoKyotoDataF.bin            (current bin)
  gaia/ExoKyotoData/data/past/                                (archived bins)
  gaia/ExoKyotoData/internal/latest/ExoKyotoDataF*.xlsx       (current xlsx)
  gaia/ExoKyotoData/internal/historical/                      (archived xlsx)

Usage:
  python3 exokyoto_update_pipeline.py \
      --in-xlsx ExoKyotoDataF20260519N_Cleaned_FullPapers_v2.xlsx \
      --data-version 2026.05.19N-preview \
      --release-notes "Added KMT-2021-BLG-0424 b and TOI-7716 b" \
      [--skip-nasa] [--skip-ads] [--skip-dedup] [--dry-run]

Output:
  Updates `gaia/ExoKyotoData/{data,internal}/latest/`
  Writes `version.json` to current directory and prints next-step git cmds.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import struct
import subprocess
import sys
import time
import urllib.parse
import urllib.request
from datetime import datetime, timezone, timedelta
from pathlib import Path

JST = timezone(timedelta(hours=9), name="JST")

def auto_data_version() -> str:
    """JST-stamped data_version, e.g. '2026.05.19-1845-JST'.

    The stamp is part of the manifest *and* the URL filename can stay the same
    — the stamp changes every minute so two pushes on the same day will never
    collide, and a sidecar version file in the client app keeps the loop
    finite (no compile-time comparison)."""
    return datetime.now(JST).strftime("%Y.%m.%d-%H%M-JST")
from typing import Dict, List, Optional, Set, Tuple

# -----------------------------------------------------------------------------
# Paths (resolved relative to this script's location)
# -----------------------------------------------------------------------------
GAIA_ROOT          = Path(__file__).resolve().parent
EKD_ROOT           = GAIA_ROOT / "ExoKyotoData"
DATA_LATEST_DIR    = EKD_ROOT / "data"     / "latest"
DATA_PAST_DIR      = EKD_ROOT / "data"     / "past"
INT_LATEST_DIR     = EKD_ROOT / "internal" / "latest"
INT_HISTORICAL_DIR = EKD_ROOT / "internal" / "historical"
INT_HIST_DATA_DIR  = INT_HISTORICAL_DIR / "data"   # bins beyond "previous"
INT_HIST_XLSX_DIR  = INT_HISTORICAL_DIR / "xlsx"   # superseded master xlsx
BIN_NAME           = "ExoKyotoDataF.bin"
CSV_NAME           = "ExoKyotoDataF.csv"

# EKDBIN1 format
MAGIC   = b"EKDBIN1\x00"
VERSION = 1

# NASA TAP
NASA_TAP_URL = "https://exoplanetarchive.ipac.caltech.edu/TAP/sync"

# ADS
ADS_URL = "https://api.adsabs.harvard.edu/v1/search/query"

# Column zones (1-based)
COL_NAME             = 1
COL_ALTERNATE_NAMES  = 67
NASA_ZONE_START      = 120  # cols 120-153 are NASA-sourced
NASA_ZONE_END        = 153
TITLE_COL_NAME       = "Title"
AUTHORS_COL_NAME     = "Authors"
JOURNAL_COL_NAME     = "Journal"
URL_COL_NAME         = "URL"
ABSTRACT_COL_NAME    = "Abstract"


# -----------------------------------------------------------------------------
# Step 1: diff
# -----------------------------------------------------------------------------
def diff_xlsx(new_xlsx: Path, prev_xlsx: Optional[Path]) -> Tuple[Set[str], Set[str]]:
    """Return (added_names, removed_names)."""
    from openpyxl import load_workbook
    def names(p):
        wb = load_workbook(p, read_only=True, data_only=True)
        ws = wb.active
        out = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                out.add(str(row[0]).strip())
        return out
    new_names = names(new_xlsx)
    if prev_xlsx is None or not prev_xlsx.exists():
        return new_names, set()
    prev_names = names(prev_xlsx)
    return (new_names - prev_names, prev_names - new_names)


# -----------------------------------------------------------------------------
# Step 2: NASA enrichment for selected planets
# -----------------------------------------------------------------------------
NASA_FIELDS = [
    "pl_name", "hostname", "discoverymethod", "disc_year", "disc_facility",
    "pl_orbper", "pl_rade", "pl_radj", "pl_masse", "pl_massj",
    "pl_tranmid", "pl_trandep", "pl_trandur",
    "pl_orbsmax", "pl_orbeccen", "pl_orbincl",
    "st_teff", "st_rad", "st_mass", "sy_dist", "ra", "dec",
    "disc_refname",
]

def nasa_query(planet_name: str) -> Optional[dict]:
    """Returns a dict of NASA fields for the planet, or None if not found."""
    sql = f"SELECT {','.join(NASA_FIELDS)} FROM pscomppars WHERE pl_name = '{planet_name}'"
    try:
        params = urllib.parse.urlencode({"query": sql, "format": "csv"})
        req = urllib.request.Request(NASA_TAP_URL + "?" + params)
        with urllib.request.urlopen(req, timeout=20) as r:
            body = r.read().decode("utf-8", errors="replace")
        rows = list(csv.DictReader(body.splitlines()))
        return rows[0] if rows else None
    except Exception as e:
        print(f"   NASA query error for {planet_name}: {e}", file=sys.stderr)
        return None


# -----------------------------------------------------------------------------
# Step 3: ADS lookup
# -----------------------------------------------------------------------------
def ads_search_planet(planet_name: str, host: str, token: str) -> Optional[dict]:
    """Search ADS for a discovery paper of <planet_name>. Returns ADS doc or None."""
    queries = []
    if host:
        queries.append(f'abs:"{planet_name}" abs:"{host}"')
    queries.append(f'abs:"{planet_name}"')
    queries.append(f'full:"{planet_name}"')
    fl = "bibcode,title,author,abstract,year,pub,doi"
    for q in queries:
        try:
            params = urllib.parse.urlencode({"q": q, "fl": fl, "rows": 3, "sort": "date asc"})
            req = urllib.request.Request(
                ADS_URL + "?" + params,
                headers={"Authorization": f"Bearer {token}"},
            )
            with urllib.request.urlopen(req, timeout=20) as r:
                data = json.loads(r.read().decode("utf-8"))
            docs = data.get("response", {}).get("docs", [])
            if docs:
                return docs[0]
        except Exception as e:
            print(f"   ADS search error: {e}", file=sys.stderr)
        time.sleep(0.3)
    return None


def format_authors(authors_list: List[str]) -> str:
    if not authors_list: return ""
    short = []
    for a in authors_list[:3]:
        parts = a.split(",", 1)
        if len(parts) == 2:
            surname = parts[0].strip()
            firsts = parts[1].strip()
            inits = ".".join(w[0] for w in firsts.split() if w) + "."
            short.append(f"{surname}, {inits}")
        else:
            short.append(a)
    s = ", ".join(short)
    if len(authors_list) > 3: s += " et al."
    return s


def fill_paper_row(ws, row_idx: int, headers: List[str], ads_doc: dict):
    def idx(h): return headers.index(h) + 1 if h in headers else -1
    def set_(col, val):
        c = idx(col)
        if c > 0: ws.cell(row_idx, c).value = val
    title = ads_doc.get("title")
    if isinstance(title, list): title = title[0] if title else ""
    authors = format_authors(ads_doc.get("author", []) or [])
    pub = ads_doc.get("pub") or ""
    year = ads_doc.get("year") or ""
    journal = f"{pub} ({year})" if pub and year else (pub or str(year))
    bib = ads_doc.get("bibcode", "")
    url = f"https://ui.adsabs.harvard.edu/abs/{bib}/abstract" if bib else ""
    abstract = ads_doc.get("abstract") or ""
    if len(abstract) > 30000:
        abstract = abstract[:29990] + "...[truncated]"
    set_(TITLE_COL_NAME,    title or "")
    set_(AUTHORS_COL_NAME,  authors)
    set_(JOURNAL_COL_NAME,  journal)
    set_(URL_COL_NAME,      url)
    set_(ABSTRACT_COL_NAME, abstract)


def fill_nasa_row(ws, row_idx: int, headers: List[str], nasa: dict):
    """Map NASA TAP fields to NASA-zone columns (120-153)."""
    nasa_to_col = {
        "pl_name":      "nasa_pl_name",
        "hostname":     "nasa_hostname",
        "discoverymethod": "nasa_discoverymethod",
        "disc_year":    "nasa_disc_year",
        "disc_facility":"nasa_disc_facility",
        "pl_orbper":    "nasa_pl_orbper",
        "pl_rade":      "nasa_pl_rade",
        "pl_radj":      "nasa_pl_radj",
        "pl_masse":     "nasa_pl_masse",
        "pl_massj":     "nasa_pl_massj",
        "pl_tranmid":   "nasa_pl_tranmid",
        "pl_trandep":   "nasa_pl_trandep",
        "pl_trandur":   "nasa_pl_trandur",
        "pl_orbsmax":   "nasa_pl_orbsmax",
        "pl_orbeccen":  "nasa_pl_orbeccen",
        "pl_orbincl":   "nasa_pl_orbincl",
        "st_teff":      "nasa_st_teff",
        "st_rad":       "nasa_st_rad",
        "st_mass":      "nasa_st_mass",
        "sy_dist":      "nasa_sy_dist",
        "ra":           "nasa_ra",
        "dec":          "nasa_dec",
        "disc_refname": "nasa_disc_refname",
    }
    for nasa_key, col_name in nasa_to_col.items():
        if col_name in headers:
            v = nasa.get(nasa_key, "")
            if v not in (None, ""):
                ws.cell(row_idx, headers.index(col_name) + 1).value = v


# -----------------------------------------------------------------------------
# Step 4: dedup (delegated)
# -----------------------------------------------------------------------------
def run_dedup_check(xlsx_path: Path) -> None:
    """Invoke check_exokyotodataf_duplicates.py as a subprocess (for visibility)."""
    script = GAIA_ROOT / "check_exokyotodataf_duplicates.py"
    if not script.exists():
        print(f"  [dedup] script not found at {script}, skipping")
        return
    print(f"  [dedup] running {script.name} on {xlsx_path.name}")
    rc = subprocess.run(
        [sys.executable, str(script), "--in-xlsx", str(xlsx_path)],
        cwd=str(GAIA_ROOT),
    ).returncode
    print(f"  [dedup] exit code = {rc}")


# -----------------------------------------------------------------------------
# Step 5-7: xlsx -> CSV -> bin, archive old
# -----------------------------------------------------------------------------
def xlsx_to_csv(xlsx_path: Path, csv_path: Path) -> int:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        n = 0
        for row in ws.iter_rows(values_only=True):
            w.writerow(["" if v is None else v for v in row])
            n += 1
    return n


def csv_to_bin(csv_path: Path, bin_path: Path) -> int:
    with csv_path.open("r", encoding="utf-8", newline="") as f:
        lines = f.read().splitlines()
    with bin_path.open("wb") as out:
        out.write(MAGIC)
        out.write(struct.pack("<I", VERSION))
        out.write(struct.pack("<I", len(lines)))
        for line in lines:
            data = line.encode("utf-8")
            out.write(struct.pack("<I", len(data)))
            out.write(data)
    return len(lines)


def rotate_bins(stamp: str, dry_run: bool = False) -> Optional[Path]:
    """Two-step rotation to keep data/past/ slim (only 1 previous bin):
       1. Any existing data/past/ExoKyotoDataF*.bin → internal/historical/data/
       2. Current data/latest/ExoKyotoDataF.bin   → data/past/ExoKyotoDataF<stamp>.bin

       Result: data/past/ holds at most one bin (the immediately previous),
               full history accumulates in internal/historical/data/.
       Returns the path of the new data/past/ entry (or None if nothing to rotate)."""
    INT_HIST_DATA_DIR.mkdir(parents=True, exist_ok=True)
    DATA_PAST_DIR.mkdir(parents=True, exist_ok=True)

    # Step 1: evict whatever is currently in data/past/ → internal/historical/data/
    moved_to_hist = []
    for old_past in sorted(DATA_PAST_DIR.glob("ExoKyotoDataF*.bin")):
        dest = INT_HIST_DATA_DIR / old_past.name
        n = 2
        while dest.exists():
            dest = INT_HIST_DATA_DIR / (old_past.stem + f"-{n}" + old_past.suffix)
            n += 1
        if not dry_run:
            shutil.move(str(old_past), str(dest))
        moved_to_hist.append(dest.name)
    if moved_to_hist:
        tag = "[dry-run] would rotate" if dry_run else "[rotate]"
        print(f"  {tag} evicted from data/past/ → internal/historical/data/: {moved_to_hist}")

    # Step 2: current latest → data/past/<stamp>
    current = DATA_LATEST_DIR / BIN_NAME
    if not current.exists():
        print(f"  [rotate] no current latest bin at {current} — nothing to demote")
        return None
    dest = DATA_PAST_DIR / f"ExoKyotoDataF{stamp}.bin"
    n = 2
    while dest.exists():
        dest = DATA_PAST_DIR / f"ExoKyotoDataF{stamp}-{n}.bin"
        n += 1
    if not dry_run:
        shutil.move(str(current), str(dest))
    tag = "[dry-run] would rotate" if dry_run else "[rotate]"
    print(f"  {tag} {current.name} → data/past/{dest.name}")
    return dest


def archive_old_master_xlsx() -> List[str]:
    """If internal/latest/ has more than one *_Cleaned_FullPapers*.xlsx (excluding
    sidecar reports), move all but the newest to internal/historical/xlsx/.
    Keeps internal/latest/ lean: master xlsx (just one) + bin + csv + sidecar reports."""
    INT_HIST_XLSX_DIR.mkdir(parents=True, exist_ok=True)
    cands = [p for p in INT_LATEST_DIR.glob("ExoKyotoDataF*FullPapers*.xlsx")
             if not re.search(r"(NasaTransitCoverage|ChangeLog|Report)", p.name)]
    if len(cands) <= 1:
        return []
    cands_sorted = sorted(cands)
    older = cands_sorted[:-1]   # all but newest
    moved = []
    for p in older:
        dest = INT_HIST_XLSX_DIR / p.name
        n = 2
        while dest.exists():
            dest = INT_HIST_XLSX_DIR / (p.stem + f"-{n}" + p.suffix)
            n += 1
        shutil.move(str(p), str(dest))
        moved.append(dest.name)
    if moved:
        print(f"  [archive xlsx] {moved} → internal/historical/xlsx/")
    return moved


# -----------------------------------------------------------------------------
# Step 8: version.json
# -----------------------------------------------------------------------------
def write_version_json(out_path: Path, data_version: str, notes: str,
                       binary_url: str) -> None:
    now_jst = datetime.now(JST)
    now_utc = now_jst.astimezone(timezone.utc)
    payload = {
        "project": "ExoKyoto3D",
        "repository": "ExoKyoto3D-Data",
        "data_version":         data_version,
        "release_date":         now_jst.strftime("%Y-%m-%d"),
        "release_time_jst":     now_jst.strftime("%H:%M:%S"),
        "release_datetime_jst": now_jst.strftime("%Y-%m-%d %H:%M:%S+09:00"),
        "release_datetime_utc": now_utc.strftime("%Y-%m-%d %H:%M:%SZ"),
        "status":     "Active",
        "data_file":  BIN_NAME,
        "data_path":  f"data/latest/{BIN_NAME}",
        "binary_url": binary_url,
        "notes":      notes,
    }
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n",
                        encoding="utf-8")
    print(f"  [version.json] wrote {out_path}  data_version={data_version}")


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--in-xlsx",      required=True, help="New xlsx (the data update)")
    ap.add_argument("--prev-xlsx",    default=None,  help="Previous xlsx (default: internal/latest/*)")
    ap.add_argument("--data-version", default=None,
        help="New version stamp. If omitted, auto-generated from current JST: "
             "YYYY.MM.DD-HHMM-JST (e.g. 2026.05.19-1845-JST)")
    ap.add_argument("--release-notes",default="",    help="Short release note (goes into version.json)")
    ap.add_argument("--archive-stamp",default=None,  help="Archive name stamp for old bin (default: YYYYMMDDNN)")
    ap.add_argument("--binary-url",
        default="https://raw.githubusercontent.com/YosukeAlex/ExoKyoto3D-Data/main/data/latest/ExoKyotoDataF.bin",
        help="The download URL written into version.json")
    ap.add_argument("--ads-token",    default=os.environ.get("ADS_TOKEN", ""), help="ADS API token (or env ADS_TOKEN)")
    ap.add_argument("--skip-diff",    action="store_true")
    ap.add_argument("--skip-nasa",    action="store_true")
    ap.add_argument("--skip-ads",     action="store_true")
    ap.add_argument("--skip-dedup",   action="store_true")
    ap.add_argument("--skip-bin",     action="store_true")
    ap.add_argument("--skip-archive", action="store_true")
    ap.add_argument("--skip-version-json", action="store_true")
    ap.add_argument("--dry-run",      action="store_true", help="Run all checks but do not write outputs")
    args = ap.parse_args()

    src_xlsx = Path(args.in_xlsx).resolve()
    if not src_xlsx.is_file():
        sys.exit(f"ERROR: in-xlsx not found: {src_xlsx}")

    # Auto-generate JST-stamped data_version if user didn't supply one
    if not args.data_version:
        args.data_version = auto_data_version()
        print(f"[init] auto data_version: {args.data_version}")

    # Determine prev_xlsx BEFORE we copy the new one in (else new becomes "previous")
    if args.prev_xlsx:
        prev_xlsx = Path(args.prev_xlsx).resolve()
    else:
        cands = [p for p in INT_LATEST_DIR.glob("ExoKyotoDataF*FullPapers*.xlsx")
                 if not re.search(r"(NasaTransitCoverage|ChangeLog|Report)", p.name)]
        cands = [p for p in cands if p.resolve() != src_xlsx.resolve()
                 and p.name != src_xlsx.name]
        prev_xlsx = sorted(cands)[-1] if cands else None

    # Copy source xlsx to internal/latest/ FIRST. All subsequent NASA/ADS write
    # ops act on this working copy — avoids Excel-lock issues on the source and
    # leaves the user's file pristine.
    INT_LATEST_DIR.mkdir(parents=True, exist_ok=True)
    new_xlsx = INT_LATEST_DIR / src_xlsx.name
    if src_xlsx.resolve() != new_xlsx.resolve():
        if not args.dry_run:
            shutil.copy2(src_xlsx, new_xlsx)
        print(f"[init] working copy: {new_xlsx}")
    else:
        print(f"[init] source already in place: {new_xlsx}")

    # ---- 1. diff ----
    print(f"\n=== [1/8] DIFF ===")
    if args.skip_diff or prev_xlsx is None:
        added, removed = set(), set()
        print(f"  skipped (or no prev)")
    else:
        added, removed = diff_xlsx(new_xlsx, prev_xlsx)
        print(f"  prev : {prev_xlsx.name}")
        print(f"  new  : {new_xlsx.name}")
        print(f"  ADDED   ({len(added)}): {sorted(added)[:10]}{' ...' if len(added)>10 else ''}")
        print(f"  REMOVED ({len(removed)}): {sorted(removed)[:10]}{' ...' if len(removed)>10 else ''}")

    # ---- 2. NASA enrichment ----
    print(f"\n=== [2/8] NASA Archive enrichment ===")
    if args.skip_nasa or not added:
        print(f"  skipped")
        nasa_results = {}
    else:
        from openpyxl import load_workbook
        wb = load_workbook(new_xlsx)  # writable
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        nasa_results = {}
        # Find row indices for added planets
        name_to_row = {}
        for r in range(2, ws.max_row + 1):
            n = ws.cell(r, COL_NAME).value
            if n: name_to_row[str(n).strip()] = r
        for planet in sorted(added):
            r = name_to_row.get(planet)
            if not r: continue
            print(f"  query: {planet}")
            res = nasa_query(planet)
            nasa_results[planet] = res
            if res:
                print(f"    FOUND in NASA — filling cols")
                fill_nasa_row(ws, r, headers, res)
            else:
                print(f"    not in NASA (ExoKyoto-original)")
        if not args.dry_run:
            wb.save(new_xlsx)
            print(f"  saved NASA fills back into {new_xlsx.name}")

    # ---- 3. ADS ----
    print(f"\n=== [3/8] ADS paper lookup ===")
    if args.skip_ads or not added:
        print(f"  skipped")
    elif not args.ads_token:
        print(f"  no ADS_TOKEN — skipping")
    else:
        from openpyxl import load_workbook
        wb = load_workbook(new_xlsx)
        ws = wb.active
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        name_to_row = {}
        for r in range(2, ws.max_row + 1):
            n = ws.cell(r, COL_NAME).value
            if n: name_to_row[str(n).strip()] = r
        for planet in sorted(added):
            r = name_to_row.get(planet)
            if not r: continue
            existing_title = ws.cell(r, headers.index(TITLE_COL_NAME) + 1).value if TITLE_COL_NAME in headers else None
            if existing_title and str(existing_title).strip():
                print(f"  {planet}: paper already filled, skip")
                continue
            host = ws.cell(r, 69).value if ws.max_column >= 69 else ""  # star_name col
            host_str = str(host).strip() if host else ""
            print(f"  search: {planet}  (host={host_str})")
            doc = ads_search_planet(planet, host_str, args.ads_token)
            if doc:
                print(f"    FOUND bibcode={doc.get('bibcode','?')}")
                fill_paper_row(ws, r, headers, doc)
            else:
                print(f"    no ADS hit")
            time.sleep(0.4)
        if not args.dry_run:
            wb.save(new_xlsx)
            print(f"  saved ADS fills back into {new_xlsx.name}")

    # ---- 4. dedup ----
    print(f"\n=== [4/8] Duplicate check ===")
    if args.skip_dedup:
        print(f"  skipped")
    else:
        run_dedup_check(new_xlsx)

    # ---- 5-6. xlsx -> CSV -> bin ----
    print(f"\n=== [5-6/8] Generate CSV + bin ===")
    DATA_LATEST_DIR.mkdir(parents=True, exist_ok=True)
    INT_LATEST_DIR.mkdir(parents=True, exist_ok=True)
    csv_internal = INT_LATEST_DIR / CSV_NAME
    bin_internal = INT_LATEST_DIR / BIN_NAME
    bin_canonical = DATA_LATEST_DIR / BIN_NAME

    if args.dry_run:
        print(f"  [dry-run] would write CSV: {csv_internal}")
    else:
        n_rows = xlsx_to_csv(new_xlsx, csv_internal)
        print(f"  CSV : {csv_internal}  ({n_rows} lines)")
    if args.skip_bin:
        print(f"  bin skipped")
    else:
        # Rotate: existing past→historical, current latest→past (one previous only)
        if not args.skip_archive:
            stamp = args.archive_stamp or datetime.now().strftime("%Y%m%d") + "01"
            rotate_bins(stamp, dry_run=args.dry_run)
        if args.dry_run:
            print(f"  [dry-run] would write bin: {bin_internal} and copy to {bin_canonical}")
        else:
            n_lines = csv_to_bin(csv_internal, bin_internal)
            print(f"  bin (internal/latest): {bin_internal}  ({n_lines} lines, {bin_internal.stat().st_size} bytes)")
            shutil.copy2(bin_internal, bin_canonical)
            print(f"  bin (data/latest)    : {bin_canonical}  (copy of above)")

    # ---- 7. tidy internal/latest (working xlsx already in place) ----
    print(f"\n=== [7/8] Demote older master xlsx ===")
    if not args.dry_run:
        archive_old_master_xlsx()

    # ---- 8. version.json ----
    print(f"\n=== [8/8] version.json ===")
    if args.skip_version_json:
        print(f"  skipped")
    else:
        notes = args.release_notes
        if added and not notes:
            notes = f"Added: {', '.join(sorted(added))}"
        vj = EKD_ROOT / "version.json"
        if not args.dry_run:
            write_version_json(vj, args.data_version, notes, args.binary_url)
        else:
            print(f"  [dry-run] would write {vj}")

    # ---- Final: git push instructions ----
    print(f"\n=== NEXT STEP: push to GitHub (manual) ===")
    print(f"  # 1) cd to your local clone of ExoKyoto3D-Data")
    print(f"  # 2) Replace these two files:")
    print(f"  #     <clone>/version.json                          <-  {EKD_ROOT}/version.json")
    print(f"  #     <clone>/data/latest/ExoKyotoDataF.bin         <-  {bin_canonical}")
    print(f"  #")
    print(f"  # (Tip: if your local clone IS {EKD_ROOT} itself, just")
    print(f"  #       'cd {EKD_ROOT} && git add -A && git commit && git push')")
    print(f"  # 3) git add -A && git commit -m 'data: {args.data_version}' && git push")
    print(f"\nDone.\n")
    return 0


if __name__ == "__main__":
    sys.exit(main())
